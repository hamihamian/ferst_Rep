# import xlsxwriter


# # Create an new Excel file and add a worksheet.
# workbook = xlsxwriter.Workbook('kdf4.xlsx')
# worksheet = workbook.add_worksheet("hamian")
# worksheet1 = workbook.add_worksheet()

# # Widen the first column to make the text clearer.
# worksheet.set_column('A:A', 20)
# worksheet1.set_column('B:B', 20)

# # Add a bold format to use to highlight cells.
# bold = workbook.add_format({'bold': True})

# # Write some simple text.
# worksheet.write('A1', 'Hello')
# worksheet1.write('B1', 'Hello')
# # Text with formatting.
# worksheet.write('A2', 'World', bold)

# # Write some numbers, with row/column notation.
# worksheet.write(2, 0, 123)
# worksheet.write(3, 0, 123.456)
# worksheet1.write(3, 0, 123.456)
# # Insert an image.
# worksheet.insert_image('B5', 'logo.png')

# workbook.close()
import openpyxl

book = openpyxl.Workbook()
book.create_sheet('Sample')
# Acquire a sheet by its name
sheet = book.get_sheet_by_name('Sample')
# Merging first 3 columns of 1st row
sheet.merge_cells('A1:C1')
# Writing to sheet
sheet.cell(row=1, column=1).value = 'sample'
book.save('kdf4.xlsx')

        
